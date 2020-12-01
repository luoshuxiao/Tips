import os
import openpyxl

i = 0


def create_execl(file_path):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = '资料详情'
    worksheet.cell(1, 1, '目录')
    worksheet.cell(1, 2, '内容')
    worksheet.cell(1, 3, '用途')
    border_thin = openpyxl.styles.Side(border_style='thin', color=openpyxl.styles.colors.BLACK)
    worksheet.column_dimensions['A'].width = 100
    worksheet.column_dimensions['B'].width = 100
    worksheet.column_dimensions['C'].width = 100
    for row in worksheet.rows:
        for cell in row:
            cell.border = openpyxl.styles.Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    path_file(file_path, worksheet)
    workbook.save(r'C:\Users\luoshuxiao\Desktop\个人资料.xlsx')


def path_file(file_path, worksheet):
    global i
    if os.path.isdir(file_path):
        worksheet.cell(row=i + 2, column=1, value=file_path)
        i += 1
        for file in os.listdir(file_path):
            path_file(file_path+file+'\\', worksheet)
    else:
        worksheet.cell(row=i + 2, column=1, value=file_path)
        i += 1
        print("第%d行" % i)


if __name__ == "__main__":
    create_execl("F:\\luoshuxiao\\")

