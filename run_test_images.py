# -*- coding: utf-8 -*-

"""
交通违法筛查平台测试脚本（把测试集数据放在img文件夹下，结果输出到json文件中，以json数据保存）
（跑测试集，测试封装逻辑的输出结果，脱离ice框架单独运行，节省时间，如果逻辑代码有更新，该测试脚本文件中的代码也需要更新）
测试集的文件目录格式如下：
   ./img/1019/510100000000A9D159/0/4/A49LR1_20190410_1710403/510100000000A9D15919041017104030000000000000000000031_A49LR1_1019.jpg
   ./img/1019/510100000000A9D159/1/A49LR1_20190410_1710403/510100000000A9D15919041017104030000000000000000000031_B57LS1_1019.jpg
"""
import itertools
import os
import cv2
import time
import json
import difflib
import logging
import datetime
import openpyxl
import threading
import traceback
import numpy as np

from collections import defaultdict
from collections import OrderedDict
from configparser import ConfigParser
from logging.config import fileConfig

from violation_1019.methods import Violation1019
from violation_1208.methods import Violation1208
from violation_1211.methods import Violation1211
from violation_1345.methods import Violation1345
from violation_1625.methods import Violation1625
from vehicle_re_recognition.methods import VehicleReid
from day_night_detection.methods import DayNightDetection
from traffic_light_detection.methods import LightDetection
from vehicle_detection_m2det.methods import VehicleDetection
from vehicle_type_detection.methods import VehicleTypeDetection
from license_plate_detection.methods import LicensePlateDetection
from license_plate_recognition.methods import LicensePlateRecognition
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

fileConfig('logging.ini')
logger = logging.getLogger(__name__)

cfg = ConfigParser()
cfg.read("setting.cfg")


#  生成算法对象的封装函数
def create_algorithm_object(cfg_info, cfg_batch_size, algorithm_class):
    detection = cfg.get('gpus', cfg_info).split(',')
    detections = [int(i) for i in detection]
    return algorithm_class(gpus=detections, batch_size=int(cfg.get('gpus', cfg_batch_size)))


# 白天黑夜检测算法对象
blur = DayNightDetection(brightness_threshold=int(cfg.get('threshold', 'percent')))
# 车辆识别算法对象
vehicle = create_algorithm_object('vehicle', 'batch_size_vehicle', VehicleDetection)
# 车牌检测算法对象
license_detection = create_algorithm_object('license_detection', 'batch_size_detection', LicensePlateDetection)
# 车牌识别算法对象
license_recognition = create_algorithm_object('license_recognition', 'batch_size_recognition', LicensePlateRecognition)
# 车辆类型算法对象
vehicle_type = create_algorithm_object('vehicle_type', 'vehicle_type_batch_size', VehicleTypeDetection)


def point_info_data():
    """
    读取本地服务器中的点位信息文件，组装数据
    :return: 系统中的点位信息数据和点位信息代号（dict, set）
    """
    points = set()  # 点位信息
    finally_point_dict = OrderedDict()  # 最终的点位信息对应数据（所有违法类型）
    point_dir = os.listdir('./config')
    for point_file in point_dir:
        point_dict = OrderedDict()  # 点位信息对应数据(单个违法类型)
        json_list = os.listdir('./config/' + point_file)
        for json_info in json_list:
            try:
                with open('./config/' + point_file + '/' + json_info, 'r') as f:
                    data_json = json.loads(f.read())
                    # 将json数据中的points中的所有数据转为int
                    for shape in data_json['shapes']:
                        shape['points'] = list(np.array(shape["points"]).astype(int))
                    # 截取点位信息文件的前23位作为点位信息josn数据对应的key(其中以1211_1625结尾的也一样，1211/1625共用)
                    point_dict[json_info.split('.')[0][:23]] = data_json
            except Exception as e:
                logger.error(e)
                logger.error(traceback.print_exc())
                logger.error(traceback.format_exc())
                continue
            points.add(json_info.split('.')[0][:23])
        finally_point_dict[point_file] = point_dict
    return finally_point_dict, points


# 加载点位信息数据（dict）和点位信息(set)到内存空间
memory_point_data, memory_point_info = point_info_data()


def transport_letter(num):
    """
    将execl中第num列转成xlsx表格的字母类型的列
    :param num: 第几列（int）
    :return: 字母型的列代号(str)
    """
    words = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    return words[num - 1] if num <= 26 else words[num // 26 - 1] + words[num % 26 - 1]


def create_sheet(type_list, worksheet, other):
    """
    创建表，并设置样式、文字等
    :param type_list: 违法类型列表 （list）
    :param worksheet: 表对象 (object)
    :param other: 是否添加 其他违法类型 到违法类型列表的判断依据(0 or 1)
    :return:
    """
    worksheet['A3'] = '点位信息'
    #  写入ill_list中的违法类型
    for index, value in enumerate(type_list + ['其他违法类型'] if other else type_list):
        start_col = 2 + index * 6
        end_col = start_col + 5
        worksheet.merge_cells('%s2:%s2' % (transport_letter(start_col), transport_letter(end_col)))
        worksheet.cell(2, 2 + index * 6, value)
        worksheet.cell(3, 2 + index * 6, '1')
        worksheet.cell(3, 3 + index * 6, '2')
        worksheet.cell(3, 4 + index * 6, '0(1)')
        worksheet.cell(3, 5 + index * 6, '0(2)')
        worksheet.cell(3, 6 + index * 6, '0(3)')
        worksheet.cell(3, 7 + index * 6, '0(4)')

    # 设置表格中单元格样式：
    worksheet.row_dimensions[1].height = 50  # 第一行行高
    worksheet.column_dimensions['A'].width = 36  # 第一列列宽
    worksheet['A1'].font = Font(name='宋体', bold=True, size=24)  # 标题（第一行）字体
    # 将第一行所有单元格合并
    worksheet.merge_cells('A1:%s1' % transport_letter(1 + len(type_list) * 6))
    border_thin = Side(border_style='thin', color=openpyxl.styles.colors.BLACK)
    for index, row in enumerate(worksheet.rows):
        # 除开第一行，每行第一列设置字体
        if index:
            row[0].font = Font(name='宋体', bold=True, size=16)
        # 将所有单元格水平、垂直居中
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            #  除了第一行意外所有单元格加边框
            if index:
                cell.border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
            # 将所有待统计的值设默认值为0
            if not cell.value and cell.row not in [1, 2]:
                cell.value = '0'
                cell.font = Font(name='宋体', bold=True, size=10)
    for col in worksheet.columns:
        col[1].font = Font(name='宋体', bold=True, size=16)
        col[2].font = Font(name='宋体', bold=True, size=16)
    worksheet.freeze_panes = 'B4'  # 冻结单元格B4 上面的行和左边的列（滚动时一直显示）


def create_execl(ill_list, spe_ill_list):
    """
    创建统计服务器处理图片情况的execl表
    :param ill_list:  一般性违法类型的列表（list）
    :param spe_ill_list:  非一般性违法类型的列表（list）
    :return:
    """
    workbook = openpyxl.Workbook()
    worksheet = workbook.active  # 第一张表

    worksheet.title = '系统已标记点位统计'
    worksheet['A1'] = '算法服务器历史数据处理统计表 -- 一般性违法类型（算法需要相应的点位json文件）'
    #  写入点位信息
    for index, item in enumerate(memory_point_info):
        worksheet['A' + str(index+4)] = item
    create_sheet(ill_list, worksheet, 0)

    workbook.create_sheet(title='系统未标记点位统计', index=1)  # 第二张表
    worksheet1 = workbook['系统未标记点位统计']
    worksheet1.column_dimensions['A'].width = 36
    worksheet1.cell(1, 1).value = '未标记模板的点位'
    worksheet1.cell(1, 1).font = Font(name='宋体', bold=True, size=16)
    worksheet1.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')

    workbook.create_sheet(title='非一般性违法统计', index=2)  # 第三张表
    worksheet2 = workbook['非一般性违法统计']
    worksheet2.column_dimensions['A'].width = 36
    worksheet2['A1'] = '算法服务器历史数据处理统计表 -- 非一般性违法类型'
    create_sheet(spe_ill_list, worksheet2, 1)
    workbook.save('./log/img_statistics.xlsx')


def update_sheet(worksheet, ill_type, point_info, result, reason):
    """
    更新表中的统计数据
    :param worksheet: 表对象
    :param ill_type: 违法类型
    :param point_info: 点位信息
    :param result: 处理结果 0，1，2（int）
    :param reason: 原因
    :return:
    """
    # 将处理图片的结果统计到excel表中
    target_row = 0  # 要更新统计数量的目标行（点位信息对应excel中的行）
    for index, item in enumerate(worksheet.columns):
        # 取第一列数据匹配点位信息（点位信息）
        if not index:
            for info in item:
                if info.value == point_info:
                    # 将匹配到的点位信息所在的行保存
                    target_row = info.row
                    break
        if item[1].value == ill_type:
            #  不合格结果对应相应违法类型的tar_column列
            if not result:
                tar_column = item[1].column + reason + 1
            #  合格结果对应相应违法类型的tar_column列
            else:
                tar_column = item[1].column + result - 1
            break
    # 如果没找到匹配的违法类型，统计到其他违法类型（即最后6列）
    else:
        if not result:
            tar_column = worksheet.max_column + reason - 4
        else:
            tar_column = worksheet.max_column + result - 6
    # 将对应的单元格的统计数量加 1
    v = worksheet.cell(target_row, tar_column).value
    val = int(v) if v else 0
    val += 1
    worksheet.cell(target_row, tar_column).value = str(val)
    worksheet.cell(target_row, tar_column).alignment = Alignment(horizontal='center', vertical='center')
    worksheet.cell(target_row, tar_column).fill = PatternFill('solid', fgColor='FAF0E6')


# 创建服务器图片处理统计execl表(服务器每次重启都会删除之前的历史，重新创建统计表)
general_illegals = [i[13:] for i in cfg.options('general_illegals')]
spe_illegals = [i[13:] for i in cfg.options('spe_illegals')]
if not os.path.exists('./log/img_statistics.xlsx'):
    create_execl(general_illegals, spe_illegals)
else:
    os.remove('./log/img_statistics.xlsx')
    create_execl(general_illegals, spe_illegals)


def traffic_restriction(org_no, week_day):
    """
    判断车牌是否限行
    :param org_no: 车牌（str）
    :param week_day: 当天周几（int）
    :return: True(限行), False(不限行)
    """
    car_no = -1  # 车牌的最后一位数字
    for item in range(len(org_no) - 1, -1, -1):
        if '0' <= org_no[item] <= '9':
            car_no = str(org_no[item])
            break
    if week_day == '1' and (car_no == '1' or car_no == '6'):
        ret = True
    elif week_day == '2' and (car_no == '2' or car_no == '7'):
        ret = True
    elif week_day == '3' and (car_no == '3' or car_no == '8'):
        ret = True
    elif week_day == '4' and (car_no == '4' or car_no == '9'):
        ret = True
    elif week_day == '5' and (car_no == '5' or car_no == '0'):
        ret = True
    else:
        ret = False
    return ret


def plate_compare(target_no, reg_no):
    """
    两个车牌是否相似（匹配四个号码以上即相似，相似度大于0.8）
    :param target_no: 目标车牌号
    :param reg_no: 算法识别出的车牌号
    :return: 目标车牌和识别出的车牌的相似度
    """
    start = 1 if ord(reg_no[0]) > 256 else 0  # 去掉车牌的汉字
    seq = difflib.SequenceMatcher(None, target_no.upper(), reg_no[start:].upper())
    return seq.ratio()


def judge_state(data, judge_type):
    """
    判断图片是否全部不合格
    :param data: 图片数组数据(list)
    :param judge_type: 图片特性检测类型(int)，如黑白检测，车辆检测，车牌检测等
    :return:
    """
    # 每组图片车辆检测是否合格（所有图片都没有检测到车辆即不合格）
    if judge_type == 1:
        for value in data.values():
            if value['car_img_count'] == 0 and value['is_ok'] != 1:
                value['result'] = 0
                value['is_ok'] = 1
                value['reason'] = 3
    # 每组图片车牌检测是否合格（所有图片都没有检测到车牌即不合格）
    if judge_type == 2:
        for value in data.values():
            if value['car_plate_count'] == 0 and value['is_ok'] != 1:
                value['result'] = 0
                value['is_ok'] = 1
                value['reason'] = 3
    # 每组图片是否有识别出车牌位数大于等于4的车牌（一个大于等于4位的车牌都没有，即不合格）
    if judge_type == 3:
        for value in data.values():
            if value['licence_count'] == 0 and value['is_ok'] != 1:
                value['result'] = 0
                value['is_ok'] = 1
                value['reason'] = 3
    # 每组图片白天、黑夜的合格性检测（全为黑夜不合格）
    if judge_type == 4:
        for value in data.values():
            if len(value['images_np']) == value['night_count'] and value['is_ok'] != 1:
                value['result'] = 0
                value['is_ok'] = 1
                value['reason'] = 3


def deal_general_illegal(illegal_type, input_dict):
    """
    不同违法类型，用不同算法处理(配置文件中相应的算法配置处于激活状态)
    :param illegal_type: 违法类型代号（str）
    :param input_dict: 输入的数据（dict）
    :return: 违法判断结果（tuple or int）
    """
    if illegal_type == '1019' and int(cfg.get('general_illegals', 'illegal_type_1019')):
        ret = Violation1019().violation_detect(input_dict, VehicleReid())
        if isinstance(ret, tuple):
            if not ret[0]:
                return 0, 4
                # return 0, ret[1] + 10190  # 避免与最终输出的不合格理由1,2,3,4冲突加个10190
            return ret
        return 0  # 返回结果不是tuple，表示算法非正常输出
    if illegal_type == '1208' and int(cfg.get('general_illegals', 'illegal_type_1208')):
        ret = Violation1208().violation_detect(input_dict, VehicleReid())
        if isinstance(ret, tuple):
            if not ret[0]:
                return 0, 4
                # return 0, ret[1]
            return ret
        return 0  # 返回结果不是tuple，表示算法非正常输出
    if illegal_type == '1211' and (cfg.get('general_illegals', 'illegal_type_1211')):
        ret = Violation1211().jugdement(input_dict, vehicle_reid=VehicleReid(), light_class=LightDetection())
        if isinstance(ret, tuple):
            if not ret[0]:
                return 0, 4
                # return 0, ret[1]
            return ret
        return 0  # 返回不是tuple类型，表示算法非正常输出
    if illegal_type == '1345' and int(cfg.get('general_illegals', 'illegal_type_1345')):
        ret = Violation1345().violation_detect(input_dict)
        if isinstance(ret, tuple):
            if not ret[0]:
                return 0, 4
            return ret
        return 0  # 不是tuple类型，返回0
    if illegal_type == '1625' and int(cfg.get('general_illegals', 'illegal_type_1625')):
        ret = Violation1625().jugdement(input_dict, vehicle_reid=VehicleReid(), light_class=LightDetection())
        if isinstance(ret, tuple):
            if not ret[0]:
                return 0, 4
                # return 0, ret[1]
            return ret
        return 0  # 不是tuple类型，返回0
    #  其他违法类型，直接返回疑似合格
    else:
        return 2, 0


def check_revise_template(input_dict):
    """
    判断传入的图片的宽、搞和点位json数据中的模板是否有偏差，高度偏差则修正，宽度偏差则返回0
    :param input_dict: 一般性违法输入的组装数据（dict）
    :return: 修正后的一般性违法输入数据或者 0 （dict or 0）
    """
    template = input_dict['location']
    h, w, _ = np.shape(input_dict['img1'][0])
    if 'imageHeight' in template.keys() and 'imageWidth' in template.keys():
        if w != int(template['imageWidth']):
            # 如果图片实际宽度与点位模板宽度不一致，返回空
            logger.info('<---该组图片点位与系统中标记的点位模板存在--宽度差，可能更换了摄像头！--->')
            return 0
        # 如果图片实际高度与点位模板高度不一致，进行修正
        if h != int(template['imageHeight']):
            shapes_new = []
            for shape in template['shapes']:
                points = np.array(shape['points'], dtype=int)
                offset_h = int(h - template['imageHeight'])
                points[:, 1] += offset_h
                shape['points'] = points
                shapes_new.append(shape)
            template['shapes'] = shapes_new
            template['imageHeight'] = h
            input_dict['location'] = template
            logger.info('<---该组图片点位与系统中标记的点位模板存在--高度差，已修正该组图片传入算法的点位标记模板！--->')
    return input_dict


class IllegalRecognitionI():

    def imagesRg(self, images, target_plates, image_times, illegal_types, image_dates, points_info, target_lanes, path, data_type_ill):
        """实现ice接口函数方法"""
        data = OrderedDict()
        ret = OrderedDict()
        ret['status'], ret['message'] = 0, ''
        ret['data'], ret['reason'], ret['reg_nos'], ret['path'] = [], [], [], path
        try:
            # 应用端数据正常才进行处理，否则直接返回结果
            img_len = len(images)
            plate_len = len(target_plates)
            illegal_len = len(illegal_types)
            date_len = len(image_dates)
            time_len = len(image_times)
            point_len = len(points_info)
            lane_len = len(target_lanes)
            if img_len == plate_len == illegal_len == date_len == time_len == point_len == lane_len:
                # 对每一组图片标记编码，解析转码，匹配车牌，时间，违法类型等
                for index, images_group in enumerate(images):
                    data[index] = OrderedDict()
                    data[index]['images_np'] = []
                    for item in images_group:
                        try:
                            # 解析图片数据、转码
                            img_np = cv2.imdecode(np.fromstring(item, np.uint8),
                                                  cv2.IMREAD_COLOR)
                        except Exception as e:
                            continue
                        if img_np is None:
                            continue
                        data[index]['images_np'].append(img_np)
                    # 对每组图片标记目标车牌号、违法类型、时间、点位信息、目标车道
                    data[index]['target_plate'] = target_plates[index]
                    data[index]['illegal_type'] = illegal_types[index]
                    data[index]['img_time'] = image_times[index]
                    data[index]['img_date'] = image_dates[index]
                    # 判断传过来的点位信息是否在算法系统中
                    # 组装点位信息格式（点位+违法类型，如果违法类型是1625，组装成1211格式，因为1211和1625点位json数据共用）
                    point_in = illegal_types[index] if illegal_types[index] != '1625' else '1211'
                    data[index]['point_info'] = points_info[index] + '_' + point_in
                    data[index]['is_point_in_system'] = 1 if data[index]['point_info'] in memory_point_info else 0
                    data[index]['target_lane'] = target_lanes[index]
                    # 初始化图片数据的相关参数
                    # result: 标明当前组图片是否合格 0不合格  1合格 2疑似合格(能匹配点位信息，默认为1，不能则默认为2)
                    data[index]['result'] = 1 if data[index]['point_info'] in memory_point_info else 2
                    data[index]['is_ok'] = 0  # 1表明当前组图片已经被判断完成；0表示未判断完成
                    data[index]['max_ratio'] = -1  # 记录识别出的车牌与目标车牌的最大相似度
                    data[index]['reg_no'] = []  # 当违法编号为13442(尾号限行)的时候才会有相应的车牌数据
                    data[index]['reason'] = 0  # 不合格原因(1：车牌模糊，2：车牌识别错误，3：其他原因，4：不属于任何违法行为)
                    data[index]['car_img_count'] = 0  # 该组检测到车辆的图片的数量
                    data[index]['night_count'] = 0  # 该组检测到黑夜图片数量
                    data[index]['car_plate_count'] = 0  # 该组所有图片检测到的车牌总数量
                    data[index]['licence_count'] = 0  # 该组识别出的车牌位数大于等于 4 的车牌数量
                # 判断应用端各类数据合格性，并做标记
                for group, images_group in data.items():
                    if len(images_group['images_np']) > 3:
                        logging.info('第%d组图片传入的图片数量超过3张，判断为疑似合格！' % (group+1))
                        images_group['result'] = 2
                        images_group['is_ok'] = 1
                        # images_group['message'] = "该组图片传入的图片数量超过3张，判断为疑似合格！"
                    if not images_group['images_np']:
                        logging.info('第%d组图片传入的图片数据有误' % (group+1))
                        images_group['result'] = 2
                        images_group['is_ok'] = 1
                        # images_group['message'] = "该组图片传入的图片数据有误"
                    if not images_group['target_plate']:
                        logging.info('第%d组图片传入的目标车牌为空' % (group+1))
                        images_group['result'] = 2
                        images_group['is_ok'] = 1
                        # images_group['message'] = "该组图片传入的目标车牌为空"
                    if not images_group['illegal_type']:
                        logging.info('第%d组图片传入的违法类型为空' % (group+1))
                        images_group['result'] = 2
                        images_group['is_ok'] = 1
                        # images_group['message'] = "该组图片传入的违法类型为空"
                    if not images_group['point_info']:
                        logging.info('第%d组图片传入的点位信息为空' % (group+1))
                        images_group['result'] = 2
                        images_group['is_ok'] = 1
                        # images_group['message'] = "该组图片传入的点位信息为空"
                    if len(str(images_group['target_lane'])) == 0:
                        logging.info('第%d组图片传入的目标车道为空' % (group+1))
                        images_group['result'] = 2
                        images_group['is_ok'] = 1
                        # images_group['message'] = "该组图片传入的目标车道为空"
                img_marks, img_data = [], []  # img_marks每张图的组号、组内编号标记; img_data所有图片数据
                # 将所有合格图片组装到一起，并标记
                for key, value in data.items():
                    if not value['is_ok']:
                        org_images = value['images_np']
                        img_tag = 0
                        for item in org_images:
                            img_marks.append([key, img_tag])
                            img_data.append(item)
                            img_tag += 1
                # 所有图片的白天黑夜识别(统计每组图片黑夜图片数量)
                s_time = time.time()
                results = blur.day_night_detection(img_data) if img_data else []
                logger.info('白天黑夜识别时间：%ss' % (time.time() - s_time))
                for index, item in enumerate(results):
                    if not item:
                        data[img_marks[index][0]]['night_count'] += 1
                # 判断各组图片是否全为黑夜图片，并做相应标记
                judge_state(data, 4)
                # 过滤掉不合格图片
                img_marks = [item for item in img_marks if data[item[0]]['result']]
                img_data = [img_data[index] for index, item in enumerate(img_marks) if data[item[0]]['result']]
                # 车辆检测
                start_time = time.time()
                results = vehicle.vehicle_detect(img_data) if img_data else []
                logger.info('车辆检测识别时间：%ss' % (time.time() - start_time))
                # 对每张图检测到的车辆进行标记
                cars_marks = []  # 每张图车辆的标记：[车辆在图片中的坐标，图片的组号，图片的组内编号，车辆在该图的编号]
                car_img = []  # 每辆车的图片截图数据
                car_tag = 0  # 车辆标签
                # 遍历拿到每一张图中所有车辆数据values和下标index
                for index, values in enumerate(results):
                    # 图片中没有车辆数据表示该图片没有检测到车辆
                    if not len(values):
                        continue
                    data[img_marks[index][0]]['car_img_count'] += 1
                    # 遍历拿到每一辆车坐标数据value
                    for value in values:
                        cars_marks.append([value, img_marks[index][0], img_marks[index][1], car_tag])
                        car_tag += 1
                        # 车辆的截图 car_data
                        car_data = img_data[index][int(value[1]):int(value[3]), int(value[0]):int(value[2])]
                        car_img.append(car_data)
                # 判断各组内所有图片能否识别到车辆，并做相应标记
                judge_state(data, 1)
                # 车牌检测
                start_time = time.time()
                results = license_detection.license_detection(car_img) if car_img else []
                logger.info('车牌检测识别时间：%ss' % (time.time()-start_time))

                licence_marks = []  # 每个车牌的标记：[坐标，图片组号，图片组内编号，车辆编号，车牌编号,车型编号]
                licence_img = []  # 车牌图片截图数据(没检测到车牌的用 [] 表示空的车牌截图数据)
                input_general = []  # 一般性违法算法输入数据的一部分(检测到的所有车辆)：[车辆坐标、车辆类型、车牌号]
                for index, value in enumerate(results):
                    mark = [None, cars_marks[index][1], cars_marks[index][2], cars_marks[index][3], index]
                    licence_data = []
                    if value:
                        mark[0] = value
                        data[cars_marks[index][1]]['car_plate_count'] += 1
                        plate_img = car_img[index][int(value[1]):int(value[3]), int(value[0]):int(value[2])]  # 车牌截图
                        licence_data = [plate_img, value[4]]  # value[4]的值：0代表车牌是一行，1代表车牌是两行
                    input_general.append([cars_marks[index][0]])
                    licence_img.append(licence_data)
                    licence_marks.append(mark)
                #  判断各组内图片是否全部未检测到车牌，全未检测到车牌，这组图片不合格
                judge_state(data, 2)
                # 车型识别（所有检测到的车辆）
                start_time = time.time()
                # 算法的参数 batch_size=1 时，可能返回空列表，导致车型结果为空，使input_dict['img1'][2]数据缺少车型数据
                results = vehicle_type.vehicle_type_detect(car_img) if car_img else []
                if car_img:
                    if not results:
                        logger.info('车型识别算法batch_size参数可能小于等于1，导致算法报错！')
                logger.info('车型识别时间：%ss' % (time.time() - start_time))

                for index, item in enumerate(results):
                    cars_marks[licence_marks[index][3]].append(item)  # 检测到的所有车辆增加车型标记
                    input_general[index].append(item)
                    licence_marks[index].append(item)
                # 车牌识别（检测到车牌的车辆）
                start_time = time.time()
                results = license_recognition.license_recogition(licence_img) if licence_img else []
                # print('车牌识别结果：', results)
                logger.info('车牌识别时间：%ss' % (time.time() - start_time))

                clear_licence_marks = []  # 识别出的车牌增加车牌和置信度标记（置信度和车牌号长度合格的车牌）
                licence_str = []  # 识别出的清晰车牌（str）
                for index, value in enumerate(results):
                    if len(value) < 2:
                        cars_marks[licence_marks[index][3]].append('000000')  # 该车辆识别出的车牌不合理，标记为车牌模糊
                        input_general[index].append('')
                    if len(value) == 2:
                        input_general[index].append(value[0])
                        #  车牌是被结果大于等于 4 ，对应组的识别到车牌的数量统计加1
                        if len(value[0]) >= 4:
                            data[licence_marks[index][1]]['licence_count'] += 1
                        if len(value[0]) < 6 or value[1] < 0.2:
                            cars_marks[licence_marks[index][3]].append('000001')
                        else:
                            cars_marks[licence_marks[index][3]].append(value)  # 车辆标记中添加车牌和置信度标记
                            clear_licence_marks.append(licence_marks[index])
                            licence_str.append(value)
                            ratio = plate_compare(data[licence_marks[index][1]]['target_plate'], value[0])
                            if ratio > data[licence_marks[index][1]]['max_ratio']:
                                data[licence_marks[index][1]]['max_ratio'] = ratio
                # 判断每组所有图片有没有识别出清晰的车牌，一个都没有则标记为不合格
                judge_state(data, 3)
                #  每组图片进行违法结果判断，并做标记
                for index, value in data.items():
                    if not value['is_ok']:
                        if value['illegal_type'] == '1340' and int(cfg.get('spe_illegals', 'illegal_type_1340')):
                            if value['max_ratio'] >= 0.8:
                                for index_licence, licence in enumerate(licence_str):
                                    if clear_licence_marks[index_licence][1] == index:
                                        #  找到目标车辆
                                        if plate_compare(value['target_plate'], licence[0]) >= 0.8:
                                            #  目标车辆是公交
                                            if clear_licence_marks[index_licence][5] == 1:
                                                value['result'] = 0
                                                value['is_ok'] = 1
                                                value['reason'] = 4
                                            #  不是公交
                                            else:
                                                if value['max_ratio'] == 1.0:
                                                    value['result'] = 1
                                                    value['is_ok'] = 1
                                                else:
                                                    value['result'] = 0
                                                    value['reason'] = 2
                                                    value['is_ok'] = 1
                                            break
                            else:
                                value['result'] = 2
                                value['is_ok'] = 1
                        elif value['illegal_type'] == '13442' and int(cfg.get('spe_illegals', 'illegal_type_13442')):
                            for index_licence, licence in enumerate(licence_str):
                                # 小轿车、货车限行判断
                                if clear_licence_marks[index_licence][1] == index:
                                    if clear_licence_marks[index_licence][5] == 2:
                                        if traffic_restriction(licence[0], value['img_time']):
                                            value['is_ok'] = 1
                                            #  如果有车牌完全匹配目标车牌，标记合格
                                            if plate_compare(value['target_plate'], licence[0]) == 1.0:
                                                value['result'] = 1
                                                value['reg_no'] = []
                                                break
                                            #  如果没有与目标车牌完全匹配的车牌，标记疑似合格，并返回识别出的限行车牌
                                            else:
                                                value['result'] = 2
                                                if licence[0] not in value['reg_no']:
                                                    value['reg_no'].append(licence[0])
                                    if clear_licence_marks[index_licence][5] == 6:
                                        if traffic_restriction(licence[0], value['img_time']):
                                            value['is_ok'] = 1
                                            value['result'] = 2
                            if not value['is_ok']:
                                value['result'] = 0
                                value['is_ok'] = 1
                                value['reason'] = 4
                        elif value['illegal_type'] == '13446' and int(cfg.get('spe_illegals', 'illegal_type_13446')):
                            for index_licence, licence in enumerate(licence_str):
                                if clear_licence_marks[index_licence][1] == index:
                                    #  能识别到货车标记为疑似合格
                                    if clear_licence_marks[index_licence][5] == 6:
                                        value['result'] = 2
                                        value['is_ok'] = 1
                                        break
                                    else:
                                        value['result'] = 0
                                        value['is_ok'] = 1
                                        value['reason'] = 4
                        else:
                            # 未识别出点位信息,标记疑似合格（已默认标记）
                            if not value['is_point_in_system']:
                                value['is_ok'] = 1
                            #  识别出点位信息
                            else:
                                # 1211/1625共用一套点位信息数据(/config文件夹下的 '1211_1625'文件夹)
                                group_ill = '1211_1625' if illegal_types[index] in ['1625', '1211'] else illegal_types[index]
                                input_dict = {
                                    "location": memory_point_data[group_ill][value['point_info']],
                                    "tarlane": value['target_lane'],
                                    "tarlicenceplate": value['target_plate'],
                                    "img1": [value['images_np'][0], value['img_date'], []] if 1 <= len(value['images_np']) else None,
                                    "img2": [value['images_np'][1], value['img_date'], []] if 2 <= len(value['images_np']) else None,
                                    "img3": [value['images_np'][2], value['img_date'], []] if 3 <= len(value['images_np']) else None,
                                }
                                img_info = []
                                #  筛选出属于index这一组的车辆
                                for index_general, value_general in enumerate(input_general):
                                    if licence_marks[index_general][1] == index:
                                        img_info.append([licence_marks[index_general][2], value_general])
                                img_dict = defaultdict(list)
                                #  根据每张图片将车辆分组
                                for item in img_info:
                                    img_dict[item[0]].append(item[1])
                                for img_index, image_info in img_dict.items():  # 组装input
                                    input_dict['img'+str(img_index+1)][2].extend(image_info)
                                result_deal = ''
                                try:
                                    #  输入组装数据图片的宽度、高度是否和系统中的标记的模板一致，
                                    #  高度不一致则修改input_dict；宽度不一致返回 0 ，标记这组图片疑似合格
                                    input_dict = check_revise_template(input_dict)
                                    #  返回给应用端提示信息,放在message里面（message里面已经有了该点位提示则不再加入）
                                    if not input_dict:
                                        mes_list = ret['message'].split('！')
                                        add_mes = 1
                                        for mes in mes_list:
                                            if value['point_info'] == mes[:23]:
                                                add_mes = 0
                                                break
                                        if add_mes:
                                            point_ill = value['point_info'][:23] if value['point_info'][-4:] != '1211' else value['point_info'][:23] + '_1625'
                                            ret['message'] += '！%s点位的图片宽度与系统中标记的模板不一致！请检查原因！' % point_ill
                                    #  根据不同的违法类型处理图片
                                    result_deal = deal_general_illegal(value['illegal_type'], input_dict) if input_dict else (2, 0)
                                except Exception as e:
                                    logger.error(e)
                                    logger.error(traceback.print_exc())
                                    logger.error(traceback.format_exc())
                                #  标记处理的结果
                                if isinstance(result_deal, tuple):
                                    value['result'] = result_deal[0]
                                    value['is_ok'] = 1
                                    value['reason'] = result_deal[1] if result_deal[1] else 0
                                else:
                                    value['result'] = 2
                                    value['is_ok'] = 1
                                    logger.info('<---算法内部出现错误，返回结果异常！--->')
                #  组装最终返回给应用端的结果 ret ,并将结果写入到excel表中
                ret['data'] = [0 for _ in range(len(data))]
                ret['reason'] = [0 for _ in range(len(data))]
                ret['reg_nos'] = [0 for _ in range(len(data))]
                #  加载execl表(如果不存在就创建，避免服务器在运行过程中，有人删除了img_statistics表造成数据不能正常处理)
                #  也实现了主动重置excel统计数据为0的功能（即删除excel表，自动从新统计结果）
                if not os.path.exists('./log/img_statistics.xlsx'):
                    create_execl(general_illegals, spe_illegals)
                workbook = openpyxl.load_workbook('./log/img_statistics.xlsx')
                worksheet = workbook.worksheets[0]
                worksheet1 = workbook.worksheets[1]
                worksheet2 = workbook.worksheets[2]
                for group, value in data.items():
                    if not value['is_ok']:
                        ret['data'][group] = 2
                        continue
                    ret['data'][group] = value['result']
                    if not value['result']:
                        ret['reason'][group] = value['reason']
                    if value['reg_no']:
                        ret['reg_nos'][group] = value['reg_no']
                    #  输出不合格理由的结果不是1,2,3,4 则不写入excel表中（避免写入excel报错，生产环境不存在这个结果，应注释掉）
                    if not value['result'] and value['reason'] not in [1, 2, 3, 4]:
                        continue
                    #  统计一般性违法处理的历史数据情况
                    if value['illegal_type'] in general_illegals:
                        # 将点位信息不存在于系统的点位统计到表中(一般性违法类型的点位信息)：
                        if not value['is_point_in_system']:
                            _point_info_general = [i.value for i in worksheet1.columns.__next__()]  # 取第一列中所有数据（点位信息）
                            if value['point_info'] not in _point_info_general:
                                worksheet1.cell(worksheet1.max_row + 1, 1).value = value['point_info']
                                worksheet1.cell(worksheet1.max_row, 1).alignment = Alignment(horizontal='center', vertical='center')
                        else:
                            update_sheet(worksheet, value['illegal_type'], value['point_info'], value['result'], value['reason'])
                    #  统计非一般性违法类型图片处理的历史数据情况
                    else:
                        _point_info = [i.value for i in worksheet2.columns.__next__()]  # 取第一列中所有数据（点位信息）
                        #  没在表中，则加进去
                        if len(value['point_info']) >= 18:
                            if value['point_info'][:18] not in _point_info:
                                cel = worksheet2.cell(worksheet2.max_row + 1, 1)
                                cel.value = value['point_info'][:18]
                                cel.alignment = Alignment(horizontal='center', vertical='center')
                                cel.font = Font(name='宋体', bold=True, size=16)
                            update_sheet(worksheet2, value['illegal_type'], value['point_info'][:18], value['result'], value['reason'])
                workbook.save('./log/img_statistics.xlsx')  # 将修改的统计信息更新到excel表中
                logger.info('总时间：%ss' % (time.time()-s_time))
            else:
                logger.info('应用端数据异常，数组长度不一致')
                ret['status'] = 1
                ret['message'] = '应用端数据异常，数组长度不一致'
        except Exception as e:
            logger.error(e)
            logger.error(traceback.print_exc())
            logger.error(traceback.format_exc())
            ret['status'] = 1
            ret['message'] += str(e)
        finally:
            logger.info('返回结果：%s' % ret)
            print(json.dumps(ret))
            load2excel(ret, data_type_ill)
            # return json.dumps(ret)


def group_img(FILE_NAME):
    """生成器(元素是每组图片的路径)"""
    for point in os.listdir(FILE_NAME):
        for result_type in os.listdir(FILE_NAME + point):
            if int(result_type):
                for group_name in os.listdir(FILE_NAME + point + '/' + result_type):
                    yield FILE_NAME + point + '/' + result_type + '/' + group_name
            else:
                for reason in os.listdir(FILE_NAME + point + '/' + result_type):
                    for group_name in os.listdir(FILE_NAME + point + '/' + result_type + '/' + reason):
                        yield FILE_NAME + point + '/' + result_type + '/' + reason + '/' + group_name


def create_data(group_name_iter):
    """
    将算法需要的图片各项数据组装
    :param group_name_iter: 多组图片的迭代器（元素是某组图片的路径）
    :return: 组装好的图片各项数据的元祖
    """
    img, target_plate, ill_type, img_date, point_info, target_lane, img_time, img_path = ([] for _ in range(8))
    for i, path in enumerate(group_name_iter):
        try:
            img.append([])
            target_plate.append('')
            ill_type.append('')
            img_date.append('')
            point_info.append('')
            target_lane.append('')
            img_time.append('')
            img_path.append(path)
            img_group = sorted(os.listdir(path), key=lambda x: x[52])  # 将图片按图片名字的第53个数字排序
            for j, p in enumerate(img_group):
                try:
                    if not j:
                        info = p.split('.')[0].split('_')
                        wek = datetime.datetime(int('20' + info[0][18:20]), int(info[0][20:22]), int(info[0][22:24])).strftime("%w")
                        ill_type[i] = info[-1]
                        target_plate[i] = info[1]
                        point_info[i] = info[0][:18]
                        img_date[i] = info[0][18:30]
                        target_lane[i] = info[0][30]
                        img_time[i] = wek
                    with open(path + '/' + p, 'rb') as f:
                        img[i].append(f.read())
                except Exception as e:
                    print(e)
                    continue
        except Exception as e:
            print(e)
            continue
    return img, target_plate, img_time, ill_type, img_date, point_info, target_lane, img_path


def create_result_excel(ill_type):
    """
    创建excel统计表
    :param ill_type: 违法类型
    :return:
    """
    work_book = openpyxl.Workbook()
    work_book.create_sheet(title='统计数量', index=0)
    work_book.create_sheet(title='合格', index=1)
    work_book.create_sheet(title='不合格原因A1（1或者2）', index=2)
    work_book.create_sheet(title='不合格原因3', index=3)
    work_book.create_sheet(title='不合格原因4', index=4)
    work_book.create_sheet(title='疑似合格', index=5)
    work_book.create_sheet(title='status状态为1', index=6)
    worksheet1 = work_book['统计数量']
    worksheet1.cell(1, 2).value = 'TRUE'
    worksheet1.cell(1, 3).value = 'FALSE'
    worksheet1.cell(2, 1).value = '合格'
    worksheet1.cell(3, 1).value = '疑似合格'
    worksheet1.cell(4, 1).value = '不合格原因A1'
    worksheet1.cell(5, 1).value = '不合格原因3'
    worksheet1.cell(6, 1).value = '不合格原因4'
    worksheet1.column_dimensions['A'].width = 15

    worksheet2 = work_book['合格']
    worksheet2.cell(1, 1).value = 'path'
    worksheet2.cell(1, 2).value = 'AI_result'
    worksheet2.column_dimensions['A'].width = 62

    worksheet3 = work_book['不合格原因A1（1或者2）']
    worksheet3.cell(1, 1).value = 'path'
    worksheet3.cell(1, 2).value = 'AI_result'
    worksheet3.column_dimensions['A'].width = 62

    # worksheet4 = work_book['不合格原因2']
    # worksheet4.cell(1, 1).value = 'path'
    # worksheet4.cell(1, 2).value = 'AI_result'
    # worksheet4.column_dimensions['A'].width = 62
    worksheet5 = work_book['不合格原因3']
    worksheet5.cell(1, 1).value = 'path'
    worksheet5.cell(1, 2).value = 'AI_result'
    worksheet5.column_dimensions['A'].width = 62

    worksheet6 = work_book['不合格原因4']
    worksheet6.cell(1, 1).value = 'path'
    worksheet6.cell(1, 2).value = 'AI_result'
    worksheet6.column_dimensions['A'].width = 62

    worksheet7 = work_book['疑似合格']
    worksheet7.cell(1, 1).value = 'path'
    worksheet7.cell(1, 2).value = 'AI_result'
    worksheet7.column_dimensions['A'].width = 62

    worksheet8 = work_book['status状态为1']
    worksheet8.cell(1, 1).value = 'path'
    worksheet8.cell(1, 2).value = 'error_result'
    worksheet8.column_dimensions['A'].width = 62

    work_book.save('./log/run_test_result_%s.xlsx' % ill_type)


def load2excel(result, ill_type):
    """
    将结果写进测试集统计excel表中
    :param result: 算法输出结果
    :param ill_type: 哪一种违法类型的数据集
    :return:
    """
    work_book = openpyxl.load_workbook('./log/run_test_result_%s.xlsx' % ill_type)
    worksheet1 = work_book.worksheets[0]  # 数量统计
    worksheet2 = work_book.worksheets[1]  # 合格
    worksheet3 = work_book.worksheets[2]  # 不合格（理由A1）
    worksheet4 = work_book.worksheets[3]  # 不合格（理由3）
    worksheet5 = work_book.worksheets[4]  # 不合格（理由4）
    worksheet6 = work_book.worksheets[5]  # 疑似合格
    worksheet7 = work_book.worksheets[6]  # error统计
    if not result['status']:
        for index, res in enumerate(result['path']):
            path_split = res.split('/')
            if path_split[4] == '0':
                if path_split[5] == 'A1':
                    if result['reason'][index] in [1, 2]:
                        val = worksheet1.cell(4, 2).value
                        val = int(val) if val else 0
                        val += 1
                        worksheet1.cell(4, 2).value = str(val)
                    else:
                        val = worksheet1.cell(4, 3).value
                        val = int(val) if val else 0
                        val += 1
                        worksheet1.cell(4, 3).value = str(val)
                        worksheet3.cell(worksheet3.max_row + 1, 1).value = res
                        worksheet3.cell(worksheet3.max_row, 2).value = '(%s,%s)' % (result['data'][index], result['reason'][index])
                        if result['data'][index] == 1:
                            worksheet3.cell(worksheet3.max_row, 2).fill = PatternFill('solid', fgColor='FAF0E6')
                else:
                    if int(path_split[5]) == result['reason'][index]:
                        if path_split[5] == '3':
                            val = worksheet1.cell(5, 2).value
                            val = int(val) if val else 0
                            val += 1
                            worksheet1.cell(5, 2).value = str(val)
                        if path_split[5] == '4':
                            val = worksheet1.cell(6, 2).value
                            val = int(val) if val else 0
                            val += 1
                            worksheet1.cell(6, 2).value = str(val)
                    else:
                        if path_split[5] == '3':
                            val = worksheet1.cell(5, 3).value
                            val = int(val) if val else 0
                            val += 1
                            worksheet1.cell(5, 3).value = str(val)
                            worksheet4.cell(worksheet4.max_row + 1, 1).value = res
                            worksheet4.cell(worksheet4.max_row, 2).value = '(%s,%s)' % (result['data'][index], result['reason'][index])
                            if result['data'][index] == 1:
                                worksheet4.cell(worksheet4.max_row, 2).fill = PatternFill('solid', fgColor='FAF0E6')
                        if path_split[5] == '4':
                            val = worksheet1.cell(6, 3).value
                            val = int(val) if val else 0
                            val += 1
                            worksheet1.cell(6, 3).value = str(val)
                            worksheet5.cell(worksheet5.max_row + 1, 1).value = res
                            worksheet5.cell(worksheet5.max_row, 2).value = '(%s,%s)' % (result['data'][index], result['reason'][index])
                            if result['data'][index] == 1:
                                worksheet5.cell(worksheet5.max_row, 2).fill = PatternFill('solid', fgColor='FAF0E6')
            if path_split[4] == '1':
                if result['data'][index] == 1:
                    val = worksheet1.cell(2, 2).value
                    val = int(val) if val else 0
                    val += 1
                    worksheet1.cell(2, 2).value = str(val)
                else:
                    val = worksheet1.cell(2, 3).value
                    val = int(val) if val else 0
                    val += 1
                    worksheet1.cell(2, 3).value = str(val)
                    worksheet2.cell(worksheet2.max_row + 1, 1).value = res
                    worksheet2.cell(worksheet2.max_row, 2).value = '(%s,%s)' % (result['data'][index], result['reason'][index])
                    if result['data'][index] == 0:
                        worksheet2.cell(worksheet2.max_row, 2).fill = PatternFill('solid', fgColor='FAF0E6')
            if path_split[4] == '2':
                if result['data'][index] == 2:
                    val = worksheet1.cell(3, 2).value
                    val = int(val) if val else 0
                    val += 1
                    worksheet1.cell(3, 2).value = str(val)
                else:
                    val = worksheet1.cell(3, 3).value
                    val = int(val) if val else 0
                    val += 1
                    worksheet1.cell(3, 3).value = str(val)
                    worksheet6.cell(worksheet6.max_row + 1, 1).value = res
                    worksheet6.cell(worksheet6.max_row, 2).value = '(%s,%s)' % (result['data'][index], result['reason'][index])
    else:
        worksheet7.cell(worksheet7.max_row + 1, 1).value = str(result['path'])
        worksheet7.cell(worksheet7.max_row, 2).value = result['message']
    work_book.save('./log/run_test_result_%s.xlsx' % ill_type)


def run_data(data_type, obj):
    if not os.path.exists('./log/run_test_result_%s.xlsx' % data_type):
        create_result_excel(data_type)
    else:
        os.remove('./log/run_test_result_%s.xlsx' % data_type)
        create_result_excel(data_type)
    img_path_iter = group_img('./img/%s/' % data_type)

    while 1:
        try:
            next(img_path_iter)
            #  生成器切片（选10组图片作为一轮去计算）
            slice_iter = itertools.islice(img_path_iter, 10)
            obj.imagesRg(*create_data(slice_iter), data_type)
            # slice_iter1 = itertools.islice(img_path_iter, 10)
            # slice_iter2 = itertools.islice(img_path_iter, 10)
            # slice_iter3 = itertools.islice(img_path_iter, 10)
            # a = threading.Thread(target=obj_img.imagesRg, args=create_data(slice_iter1))
            # b = threading.Thread(target=obj_img.imagesRg, args=create_data(slice_iter2))
            # c = threading.Thread(target=obj_img.imagesRg, args=create_data(slice_iter3))
            # a.start()
            # b.start()
            # c.start()
            # a.join()
            # b.join()
            # c.join()
        except StopIteration:
            print('%s done' % data_type)
            break


if __name__ == '__main__':
    # type_list = ['test']
    type_list = ['1208']
    # type_list = ['1208', '1019', '1625', '1211', '1345', '1340', '13442']
    obj_img = IllegalRecognitionI()
    for data in type_list:
        s = time.time()
        run_data(data, obj_img)
        print('run %s 类型，共用时间：%smin' % (data, (time.time()-s)/60))
