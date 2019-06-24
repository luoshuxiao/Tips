# -*- coding: utf-8 -*-

"""
三种方式读写修改excel文件：
一： xlrd/xlwt/xlutils -- 读/写/修改
二： openpyxl
三： pandas -- 只能实现对数据进行操作，不能实现对excel样式操作
"""

import xlrd
import xlwt
import openpyxl
import pandas as pd

from xlutils import copy
from openpyxl import styles

import pymysql
import sqlalchemy as sqla

from sqlalchemy import create_engine



# 测试用数据
province = ['北京市', '天津市', '河北省', '山西省', '内蒙古自治区', '辽宁省',
            '吉林省', '黑龙江省', '上海市', '江苏省', '浙江省', '安徽省', '福建省',
            '江西省', '山东省', '河南省', '湖北省', '湖南省', '广东省', '广西壮族自治区',
            '海南省', '重庆市', '四川省', '贵州省', '云南省', '西藏自治区', '陕西省', '甘肃省',
            '青海省', '宁夏回族自治区', '新疆维吾尔自治区']

income = ['5047.4', '3247.9', '1514.7', '1374.3', '590.7', '1499.5', '605.1', '654.9',
          '6686.0', '3104.8', '3575.1', '1184.1', '1855.5', '1441.3', '1671.5', '1022.7',
          '1199.2', '1449.6', '2906.2', '972.3', '555.7', '1309.9', '1219.5', '715.5', '441.8',
          '568.4', '848.3', '637.4', '653.3', '823.1', '254.1']

project = ['各省市', '工资性收入', '家庭经营纯收入', '财产性收入', '转移性收入', '食品', '衣着',
           '居住', '家庭设备及服务', '交通和通讯', '文教、娱乐用品及服务', '医疗保健', '其他商品及服务']


def xlrd_read_excel(filename):
    # xlrd和xlwt能处理的xls文件最大的行数为65535，用xlrd读取excel是不能对其进行操作的
    # 超过65535就会遇到错误： ValueError: row index was 65536, not allowed by .xls format
    read_e = xlrd.open_workbook(filename)  # 读取excel文件
    names = read_e.sheet_names()  # 获取所有表的表名,返回list
    print(names)
    sheet1 = read_e.sheet_by_name('Sheet1')  # 根据名字读取表
    # sheet1 = read_e.sheet_by_index(0)  # 根据索引读取表
    she_name = sheet1.name  # 获取sheet1工作表的名字
    print(she_name)
    nrows = sheet1.nrows  # 获取sheet1表的最大行
    ncols = sheet1.ncols  # 获取sheet1表的最大列
    row_value = sheet1.row_values(0)  # 返回第1行的数据，list的每个元素是str类型
    col_value = sheet1.col_values(0)  # 第一列的数据，list的每个元素是str类型
    cell_object = sheet1.cell(1, 6)
    cell_value = sheet1.cell(1, 6).value  # 获取第2行第7列单元格的值
    cell_value_row = sheet1.row(1)[6].value  # 通过行或者列定位单元格获取值

    print(f'第一行数据：{row_value}')
    print(f'第一列数据：{col_value}')
    print(ncols)
    print(nrows)
    print(f'单元格对象：{cell_object}')  # 输出结果：text:'数据量'
    print(f'单元格的值：{cell_value}')


def xlwt_write_excel(filename):
    write_e = xlwt.Workbook(encoding='utf-8',style_compression=0)  # 创建一个excel工作簿（文件）,后一个参数表示是否压缩一般不用
    test = write_e.add_sheet('test', cell_overwrite_ok=True)  # 添加一个excel表,cell_overwrite_ok表示覆盖单元格，默认False
    test2 = write_e.add_sheet('test2', cell_overwrite_ok=True)  # 添加第二个excel表
    test.write(3, 4, '测试写入数据')  # 写入数据到第4行5列单元格

    #  按行或者列批量插入数据：
    # 填入第一行
    for i in range(0, len(project)):
        test2.write(0, i, project[i])
    # 填入第一列
    for i in range(0, len(province)):
        test2.write(i + 1, 0, province[i])
    # 填入第二列
    for i in range(0, len(income)):
        test2.write(i + 1, 1, income[i])
    write_e.save(filename)  # 保存到filename文件中


def xlutils_update_excel():
    workbook = xlrd.open_workbook(r'C:\Users\luoshuxiao\Desktop\test33.xls')
    #  生成xlutils文件对象
    copy_workbook = copy.copy(workbook)
    #  拿到需要修改的表
    copy_sheet = copy_workbook.get_sheet(1)
    # 修改操作（批量操作走循环）
    copy_sheet.write(0, 0, 'changed by xlutils')
    # 保存到原excel表，也可以另起名字创建新表
    copy_workbook.save(r'C:\Users\luoshuxiao\Desktop\test33.xls')
    # xlutils官方对以下方法都有介绍：
    # xlutils.copy/xlutils.display/xlutils.filter/xlutils.margins/xlutils.styles


def openpyxl_read_excel():
    workbook = openpyxl.load_workbook(r'C:\Users\luoshuxiao\Desktop\test.xlsx')  # 读取文件，创建工作簿对象
    # she_names = workbook.get_sheet_names()  # 获取所有的工作表名字（官方已经将该方法弃用）
    she_names1 = workbook.sheetnames   # 获取所有工作表名字
    print(she_names1)   # 输出结果：['Sheet1', 'Sheet2', 'Sheet3']
    # worksheet = workbook.get_sheet_by_name('Sheet1')  # 获取表对象(官方已经将该方法弃用)
    worksheet1 = workbook['Sheet1']  # 获取表对象
    worksheet2 = workbook[she_names1[0]]  # 获取表对象
    worksheet3 = workbook.worksheets[0]  # 获取表对象
    worksheet4 = workbook.active  # 获取表对象(当前活跃表，默认第一个)
    print(worksheet1)  # 输出结果： <Worksheet "Sheet1">
    print(worksheet2)  # 输出结果： <Worksheet "Sheet1">
    print(worksheet3)  # 输出结果： <Worksheet "Sheet1">
    print(worksheet4)  # 输出结果： <Worksheet "Sheet1">
    #  获取指定工作表的属性（表名，行数，列数等）
    she_name = worksheet2.title  # 获取表名
    print(she_name)  # 输出结果： Sheet1
    rows = worksheet2.max_row  # 获取最大行数
    cols = worksheet2.max_column  # 获取最大列数
    print(rows, cols)

    # 按行、列获取表中所有数据（当然也可以通过行列下标直接循环通过cell(row,column)获取）
    # sheet.rows  获取行数据（生成器），生成器每个元素是一个元祖，保存每一行数据，元祖每个元素是对应行的每一个单元格
    # sheet.columns 获取列数据（生成器），生成器每个元素是一个元祖，保存每一列数据，元祖每个元素是对应列的每一个单元格
    for row in worksheet2.rows:   # 遍历每一行
        for cell in row:    # 遍历每一行的每一列，获取单元格对象
            cell_value = cell.value  # 获取单元格的值
            print(cell_value)
    for col in worksheet2.columns:  # 遍历每一列
        for cell in col:   # 遍历每一列的每一行，获取单元格对象
            cell_value = cell.value  # 获取单元格的值
            print(cell_value)
            print(cell.row)  # 获取单元格所在的行
            print(cell.column)  # 获取单元格所在的列
            print(cell.coordinate)  # 获取单元格的坐标 -- 'B2'

    # 获取指定行、列的数据 -- 将sheet.rows/columns生成器转换成list类型通过索引获取
    for cell in list(worksheet2.rows)[0]:
        print(cell.value)
    for cell in list(worksheet2.columns)[0]:
        print(cell.value)

    # 两种方式获取某一局部数据信息（某一行到某一行的某一列到某一列数据）
    # 方式一：将行、列迭代器转成列表进行切片获取局部数据
    for rows in list(worksheet2.rows)[0:3]:  # 获取行的切片（如果先获取列也是一样到你）
        for cell in rows[1:5]:  # 对每一行数据中的列数据切片
            print(cell.value)
    # 方式二：定位行列位置，直接动态获取局部区域每一个单元格
    # 与xlrd不同的是，openpyxl下标是从1开始，而xlrd是从0开始
    for i in range(1, 4):
        for j in range(2, 6):
            print(worksheet2.cell(row=i, column=j).value)  # cell可以直接写：cell(i,j)

    # 获取某一个单元格数据：
    c_value = worksheet2['A1'].value  # 通过excel表中的单元格定位方式定位单元格
    c_value1 = worksheet2.cell(1, 1).value  # 通过下标索引定位单元格（下标从1开始）
    print(c_value)
    print(c_value1)


def openpyxl_write_excel():
    workbook = openpyxl.Workbook()  # 创建一个Excel文件（默认utf8）
    worksheet = workbook.active  # 获取当前活跃的表（默认第一个）
    worksheet.title = 'openpyxl'  # 为该表取表名
    worksheet1 = workbook.create_sheet()  # 默认在工作簿的最后一页（传入页码参数可以插入到指定页码前）
    worksheet1.title = 'openpyxl1'
    # 边框的样式
    border_thin = Side(border_style='thin', color=openpyxl.styles.colors.BLACK)
    for row in worksheet.rows:
        for cell in row:
            # 给每一个单元格加四个边框（可以指定只加相应的上下左右边框）
            cell.border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    # 将数据一行一行或者一列一列写入excel表中（可以将整张表数据放在一个列表中写入）
    for i in range(len(project)):
        worksheet.cell(1, i+1, project[i])  # 写第一行数据，openpyxl单元格下标是从1开始的
    for i in range(len(province)):
        worksheet.cell(i+2, 1, province[i])  # 写第一列数据，i+2表示第一行数据已经写入
    for i in range(len(income)):
        worksheet.cell(i+2, 2, income[i])  # 写入第二列数据
    workbook.save(r'C:\Users\luoshuxiao\Desktop\openpyxl.xlsx')  # 将excel保存到指定位置
    # workbook.save(r'C:\Users\luoshuxiao\Desktop\openpyxl.xls')  # 虽然不能读取xls文件格式，但是openpyxl可以保存为xls文件


def openpyxl_update_excel():
    workbook = openpyxl.load_workbook(r'C:\Users\luoshuxiao\Desktop\openpyxl.xlsx')
    worksheet = workbook.worksheets[0]
    worksheet.insert_cols(1)  # 在第一列前插入一列
    # worksheet.insert_rows(1)  # 在第一行前插入一行
    update_column = [i for i in range(1, 33)]
    update_column.insert(0, '第一列')
    for index, row in enumerate(worksheet.rows):
        row[0].value = update_column[index]  # 给第一列每一行单元格赋值（循环插入时单元格个数需要匹配，不然要报下标越界）
    worksheet.cell(1, 1, '编号')  # 将第一行第一列单元格值改为：编号
    worksheet['A1'] = 'number'  # 将第一行第一列单元格值改为：number
    HongKong = [32, '香港', 2000]
    worksheet.append(HongKong)   # 添加一行数据到指定行
    worksheet.delete_rows(5, 1)  # 删除第5行后的一行数据
    worksheet.delete_cols(5, 2)  # 删除第五列后的两列数据
    # workbook.remove(worksheet)  # 删除工作表worksheet
    # del workbook[worksheet]  # 删除工作表worksheet

    # 表和单元格的样式修改：
    # 改变sheet标签颜色：
    worksheet.sheet_properties.tabColor = '660066'  # 必须是aRGB的十六进制值，不能是英文
    #  设置字体（cell的font属性）：等线24号，加粗斜体，字体颜色红色
    worksheet['A1'].font = styles.Font(name='等线', size=20, italic=True, color=styles.colors.RED, bold=True)
    #  对齐方式（cell的aligment属性）：垂直和水平居中
    worksheet['B1'].alignment = styles.Alignment(horizontal='center', vertical='center')
    #  设置行高，列宽：
    worksheet.row_dimensions[2].height = 40      # 设置第2行行高40px
    worksheet.column_dimensions['C'].width = 40      # 设置C列列宽40px
    #  合并拆分单元格：
    #  合并后的单元格在表中的定位和数据值，都以左上角单元格为准（无论是空还是有数据）
    worksheet.merge_cells('B1:G1')  # 合并一行中的几个单元格（合并后以B1为准）
    worksheet.merge_cells('A1:C3')  # 合并一个局部区域中的单元格
    #  拆分单元格也是以左上角单元格为准
    worksheet.unmerge_cells('A1:C3')  # 有值就填充到A1，拆分后的其他单元格均为空
    workbook.save(r'C:\Users\luoshuxiao\Desktop\openpyxl.xlsx')


# pandas支持很多文件类型的数据读写功能json/excel/dat/csv/等，包括html中的table，数据库中的table
def pandas_read_excel():
    xls_file = pd.ExcelFile(r'C:\Users\luoshuxiao\Desktop\爬虫项目总结.xls')   # 用pandas读取2003excel版本以上的xls文件，生成一个excel实列对象
    table = xls_file.parse('Sheet1')   # 解析出表Sheet1的内容，转换成DataFrame数据类型
    print(table.head(5))
    data = pd.read_excel(r'C:\Users\luoshuxiao\Desktop\爬虫项目总结.xls')  # 通过read_excel方法读取excel
    print(data.head(10))

    # 从mysql读取数据转为dataframe：
    db = sqla.create_engine('mysql+pymysql://root:123456@127.0.0.1/jd?charset=utf8')
    data = pd.read_sql('select * from goods', db)
    print(data)


def pandas_write_excel():
    df1 = pd.DataFrame({'data1': [1, 2, 3, 4, 'NaN', 6, 7, 8]})
    df2 = pd.DataFrame({'data2': [11, 21, 31, 41, 51, 61, 71, 81]})
    df3 = pd.DataFrame({'data3': [11, 22, 33, 44, 55, 66, 77, 88]})
    #  写入方式一：to_excel 默认的一个数据集写入一个文件中（直接传入文件名）
    df1.to_excel(r'C:\Users\luoshuxiao\Desktop\pandas_write1.xlsx', na_rep=0, index=False, header=None)
    #  写入方式二：pd.ExcelWriter创建文件对象，将多个数据集写入同一个文件
    #  生成一个excel文件对象
    writer_pd = pd.ExcelWriter(r'C:\Users\luoshuxiao\Desktop\pandas_write2.xlsx')
    #  数据df1写入到df1表中
    df1.to_excel(writer_pd, sheet_name='df1', startcol=0, index=True)
    #  数据df2写入到df1表中
    df2.to_excel(writer_pd, sheet_name='df2', startcol=4, index=False)
    #  将数据df3写入到df3表中
    df3.to_excel(writer_pd, sheet_name='df3', index_label='索引列标题')
    writer_pd.save()  # 保存

    # # pandas将datafeame写入mysql数据库:
    # path1 = r'C:\Users\luoshuxiao\Desktop\PYTHON\总结\github仓库\data-analyst\datasets\movielens\movies.dat'
    # movie_table = pd.read_csv(path1, header=None,
    #               names=['movie_id', 'title', 'genres'],
    #               sep='::', engine='python')  # 用pandas读取本地dat文件(分隔符文件)或者其他支持类型文件，生成DataFrame数据
    # connect_db = create_engine('mysql+pymysql://root:123456@127.0.0.1/movie_db?charset=utf8')  # 通过sqlalchemy创建数据库引擎
    # #  注意：如果用if_exists = 'replace'，会先删除原表，再创建新表字段，所以新建的数据表与原来不同（字段的数据类型）。
    # #   append参数不会改变原来字段的数据类型，会在原表下添加数据
    # movie_table.to_sql('movies', connect_db, if_exists='append', index=False)


def main():
    # xlrd_read_excel(r'C:\Users\luoshuxiao\Desktop\test.xls')
    # xlwt_write_excel(r'C:\Users\luoshuxiao\Desktop\test33.xls')
    # xlutils_update_excel()
    # openpyxl_read_excel()
    # openpyxl_write_excel()
    # openpyxl_update_excel()
    # pandas_read_excel()
    pandas_write_excel()


if __name__ == '__main__':
    main()
